"""
========================================================
  ALIBABA SUPPLIER FINDER – Wyszukiwarka dostawców
  Wersja 4.0 – undetected-chromedriver + Google Sheets
  Dla: Rodziców (branża torebkowa / bag hardware)

  JAK URUCHOMIĆ (raz na początku):
    pip install undetected-chromedriver openpyxl gspread google-auth

  Następnie:
    python alibaba_dostawcy.py

  WYMAGANIA:
    - Google Chrome zainstalowany na komputerze
    - ChromeDriver instaluje się AUTOMATYCZNIE

========================================================
  SETUP GOOGLE SHEETS (jednorazowo – ok. 5 minut):
========================================================

  1. Wejdź na: https://console.cloud.google.com/
  2. Stwórz nowy projekt (np. "Alibaba Finder")
  3. Włącz API:
       → "APIs & Services" → "Enable APIs"
       → Wyszukaj i włącz: "Google Sheets API"
       → Wyszukaj i włącz: "Google Drive API"
  4. Stwórz klucz:
       → "APIs & Services" → "Credentials"
       → "Create Credentials" → "Service Account"
       → Nadaj dowolną nazwę, kliknij "Done"
       → Kliknij na stworzony Service Account
       → Zakładka "Keys" → "Add Key" → "JSON"
       → Pobierz plik JSON
  5. Zmień nazwę pobranego pliku na: google_credentials.json
     i umieść go w tym samym folderze co ten skrypt.
  6. Otwórz plik JSON w notatniku, skopiuj wartość "client_email"
     (wygląda jak: xxx@xxx.iam.gserviceaccount.com)
  7. Wejdź na: https://sheets.google.com
     Stwórz nowy arkusz o nazwie: "Dostawcy Alibaba"
     Kliknij "Udostępnij" i wklej skopiowany email (uprawnienia: Edytor)
  8. Gotowe! Skrypt będzie automatycznie zapisywał dane do tego arkusza.

========================================================
"""

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date
import time
import random
import re
import sys

# Google Sheets – opcjonalna integracja
try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

# ─────────────────────────────────────────────
#  KONFIGURACJA – edytuj tutaj
# ─────────────────────────────────────────────
SEARCH_PHRASES = [
    # ── Ramki z plastikową skorupą ──
    "clutch frame with plastic shell",
    "metal purse frame with plastic box shell",
    "box clutch frame",
    "metal box purse frame",
    "clutch frame and shell",
    "clutch box frame with plastic shell",
    "metal clutch frame with plastic shell",
    "purse frame with box",
    # ── Gotowe kopertówki wieczorowe ──
    "evening clutch bag",
    "hard case box clutch evening bag",
    "wedding party clutch purse",
    "bridal clutch bag",
    "metal frame evening clutch bag",
    "rhinestone evening clutch bag",
    "crystal evening clutch bag",
    "dinner clutch bag",
    "banquet clutch bag",
]

# ── Poprzednie frazy (tymczasowo wyłączone) ──
# SEARCH_PHRASES_OLD = [
#     "Metal purse frame",
#     "Metal clutch bag frames",
#     "Woman bag clutch frame",
#     "Handbag hardware",
#     "Box purse frame",
#     "Kiss lock purse frame",
#     "Minaudiere frame metal",
#     "Evening bag frame",
#     "Metal bag frame clasp",
#     "Bag frame accessories",
#     "Metal box clutch bag",
#     "Minaudiere clutch bag",
#     "Plastic shell purse",
#     "Plastic shell box purse frame",
#     "Woman metal bag",
# ]

MIN_STARS        = 0      # minimalna ocena (0–5); 0 = zapisz wszystkich
MAX_PAGES        = 5      # ile stron na frazę (ok. 20 wyników/stronę)
DELAY_PAGE       = (3, 5) # przerwa między stronami (sekundy)
DELAY_PHRASE     = (4, 7) # przerwa między frazami (sekundy)
MAX_RUNTIME_MIN  = 720    # maksymalny czas działania w minutach (12 godzin)
CAPTCHA_WAIT_S   = 10     # ile sekund czekać na rozwiązanie CAPTCHA
OUTPUT_FILE      = f"dostawcy_alibaba_{date.today()}.xlsx"

ENRICH_PROFILES  = False  # dane są na kartach wyszukiwania – faza 2 niepotrzebna
DELAY_ENRICH     = (1, 2) # przerwa między odwiedzinami profili (tylko gdy ENRICH_PROFILES=True)

# ── Google Sheets (wypełnij po wykonaniu kroków SETUP powyżej) ──
GOOGLE_SHEETS_ENABLED = True                    # True = zapisuj do Sheets, False = tylko lokalny Excel
GOOGLE_CREDS_FILE     = "google_credentials.json"  # plik JSON z Google Cloud Console
GOOGLE_SHEET_NAME     = "Dostawcy Alibaba"     # nazwa arkusza w Google Sheets (taka sama jak w kroku 7)

# Słowa kluczowe wskazujące na CAPTCHA na stronie
CAPTCHA_SIGNALS = [
    "verify you are human", "are you a robot", "unusual traffic",
    "security check", "please verify", "captcha", "challenge",
    "access denied", "bot detection",
]


# ─────────────────────────────────────────────
#  ŚCIEŻKA DO CHROME – ustaw ręcznie jeśli auto-detect nie działa
# ─────────────────────────────────────────────
# Zostaw pusty string "" żeby wykrywać automatycznie.
# Jeśli Chrome się nie otwiera, wpisz pełną ścieżkę, np.:
#   macOS:   "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
#   Windows: r"C:\Program Files\Google\Chrome\Application\chrome.exe"
CHROME_PATH = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"


# ─────────────────────────────────────────────
#  URUCHOMIENIE PRZEGLĄDARKI
# ─────────────────────────────────────────────
import os, subprocess

def find_chrome() -> str | None:
    """Znajdź Chrome automatycznie na macOS / Windows / Linux."""
    # 1. Ręczna ścieżka z konfiguracji
    if CHROME_PATH and os.path.exists(CHROME_PATH):
        return CHROME_PATH

    # 2. macOS – szukaj przez mdfind (Spotlight)
    try:
        result = subprocess.run(
            ["mdfind", "kMDItemCFBundleIdentifier == 'com.google.Chrome'"],
            capture_output=True, text=True, timeout=5
        )
        for line in result.stdout.strip().splitlines():
            candidate = os.path.join(line.strip(), "Contents/MacOS/Google Chrome")
            if os.path.exists(candidate):
                return candidate
    except Exception:
        pass

    # 3. Standardowe ścieżki
    candidates = [
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        "/Applications/Chromium.app/Contents/MacOS/Chromium",
        os.path.expanduser("~/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        "/usr/bin/google-chrome",
        "/usr/bin/chromium-browser",
    ]
    for path in candidates:
        if os.path.exists(path):
            return path

    return None


def get_chrome_version(chrome_path: str) -> int | None:
    """Pobierz główną wersję Chrome (np. 124)."""
    try:
        result = subprocess.run(
            [chrome_path, "--version"],
            capture_output=True, text=True, timeout=5
        )
        match = re.search(r"(\d+)\.\d+", result.stdout)
        if match:
            return int(match.group(1))
    except Exception:
        pass
    return None


def create_driver() -> uc.Chrome:
    opts = uc.ChromeOptions()
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.page_load_strategy = "none"   # nie czekaj na pełne załadowanie – sami sprawdzimy

    chrome_path = find_chrome()
    version_main = None

    if chrome_path:
        print(f"  ✓ Chrome znaleziony: {chrome_path}")
        opts.binary_location = chrome_path
        version_main = get_chrome_version(chrome_path)
        if version_main:
            print(f"  ✓ Wersja Chrome: {version_main}")
    else:
        print("  ⚠️  Nie znaleziono Chrome automatycznie.")
        print("     Ustaw ścieżkę ręcznie w zmiennej CHROME_PATH na górze skryptu.")

    try:
        if version_main:
            driver = uc.Chrome(options=opts, version_main=version_main)
        else:
            driver = uc.Chrome(options=opts)
        driver.set_script_timeout(45)   # 45s na wykonanie JS (domyślnie ~30s)
        return driver
    except Exception as e:
        print(f"\n❌  Błąd uruchamiania Chrome: {e}")
        print("\nSprawdź:")
        print("  1. Czy Google Chrome jest zainstalowany?")
        print("  2. Ustaw ścieżkę ręcznie: CHROME_PATH = \"/pełna/ścieżka/do/chrome\"")
        sys.exit(1)


# ─────────────────────────────────────────────
#  WYKRYWANIE I OBSŁUGA CAPTCHA
# ─────────────────────────────────────────────
def check_and_handle_captcha(driver) -> bool:
    """Sprawdza czy jest CAPTCHA. Jeśli tak – czeka CAPTCHA_WAIT_S sekund
    aż użytkownik ją rozwiąże. Jeśli nie – pomija stronę.
    Zwraca True jeśli CAPTCHA była i została rozwiązana."""
    try:
        page_text = driver.page_source.lower()
    except Exception:
        return False

    captcha_detected = any(signal in page_text for signal in CAPTCHA_SIGNALS)
    if not captcha_detected:
        return False

    print("\n" + "=" * 55)
    print("  ⚠️   CAPTCHA WYKRYTA!")
    print("  Rozwiąż ją ręcznie w otwartej przeglądarce.")
    print(f"  Skrypt czeka {CAPTCHA_WAIT_S} sekund...")
    print("=" * 55)

    steps = max(1, CAPTCHA_WAIT_S // 2)
    for i in range(steps):
        time.sleep(2)
        try:
            page_text = driver.page_source.lower()
        except Exception:
            break
        still_captcha = any(signal in page_text for signal in CAPTCHA_SIGNALS)
        if not still_captcha:
            print("  ✅  CAPTCHA rozwiązana – kontynuuję!\n")
            return True
        remaining = CAPTCHA_WAIT_S - (i + 1) * 2
        print(f"  ⏳  Czekam... (pozostało ~{max(0, remaining)}s)")

    print("  ❌  Timeout – pomijam stronę i jadę dalej.\n")
    return False


# ─────────────────────────────────────────────
#  WYSZUKIWANIE JEDNEJ FRAZY
# ─────────────────────────────────────────────
def search_suppliers(driver, phrase: str, page: int) -> list[dict]:
    # URL produktów – sprawdzony, działa z selektorem .J-offer-wrapper
    search_q = phrase.replace(" ", "_")
    url = (
        f"https://www.alibaba.com/products/{search_q}.html"
        f"?IndexArea=product_en&page={page}"
    )

    print(f"    Ładuję: {url}")
    try:
        driver.get(url)
    except Exception:
        pass  # page_load_strategy=none – wyjątki przy get() są normalne, ignorujemy

    # Poczekaj chwilę żeby strona zaczęła się ładować
    time.sleep(4)

    # Sprawdź CAPTCHA zaraz po załadowaniu
    check_and_handle_captcha(driver)

    # Poczekaj aż pojawią się karty produktów (max 30 sekund)
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
                ".J-offer-wrapper, .organic-offer-wrapper, .fy23-search-card"))
        )
    except TimeoutException:
        # Jeszcze jedna szansa – może CAPTCHA pojawiła się po WebDriverWait
        check_and_handle_captcha(driver)
        # Spróbuj jeszcze raz przez 10 sekund
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR,
                    ".J-offer-wrapper, .organic-offer-wrapper, .fy23-search-card"))
            )
        except TimeoutException:
            print("    ⚠️  Timeout – brak wyników, pomijam stronę.")
            return []

    # Scroll
    try:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight / 2);")
        time.sleep(1.5)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
    except Exception:
        time.sleep(2)

    return parse_page(driver, phrase, url)


# ─────────────────────────────────────────────
#  PARSOWANIE STRONY
# ─────────────────────────────────────────────
def parse_page(driver, phrase: str, source_url: str) -> list[dict]:
    suppliers = []

    # Selektory kart produktów (każda karta zawiera info o dostawcy)
    card_selectors = [
        "div.J-offer-wrapper",          # klasyczny selektor (potwierdzony)
        "div.organic-offer-wrapper",
        "div.fy23-search-card",
        "div[class*='offer-wrapper']",
        "div[class*='search-card']",
        "div.list-item",
    ]

    cards = []
    for sel in card_selectors:
        try:
            found = driver.find_elements(By.CSS_SELECTOR, sel)
            if len(found) >= 2:
                cards = found
                print(f"    ✓ Znaleziono {len(cards)} kart (selektor: {sel})")
                break
        except Exception:
            continue

    if not cards:
        print("    ⚠️  Brak kart – próbuję fallback na linki /company/")
        return fallback_links(driver, phrase, source_url)

    for card in cards:
        s = extract_supplier(card, phrase, source_url)
        if s:
            suppliers.append(s)

    # Usuń duplikaty dostawców w obrębie tej samej strony
    seen = set()
    unique = []
    for s in suppliers:
        key = s["Nazwa dostawcy"].lower().strip()
        if key not in seen:
            seen.add(key)
            unique.append(s)

    return unique


def extract_supplier(card, phrase: str, source_url: str) -> dict | None:
    """Wyciąga dane dostawcy z karty produktu na Alibaba."""

    def safe_text(*selectors):
        for sel in selectors:
            try:
                el = card.find_element(By.CSS_SELECTOR, sel)
                t = el.text.strip()
                if t:
                    return t
            except Exception:   # łap TimeoutException, StaleElement itp.
                pass
        return ""

    def safe_attr(attr, *selectors):
        for sel in selectors:
            try:
                el = card.find_element(By.CSS_SELECTOR, sel)
                v = el.get_attribute(attr)
                if v:
                    return v
            except Exception:   # łap TimeoutException, StaleElement itp.
                pass
        return ""

    # ── Nazwa firmy (selektory dla kart produktów) ──
    name = safe_text(
        # nowe selektory Alibaba 2024/2025
        ".search-card-e-company",
        ".organic-gallery-offer-inner__specifications-supplier",
        "a.search-card-e-slider__title",
        # ogólne
        "[class*='company']",
        "[class*='supplier-name']",
        "[class*='store-name']",
        "h4 + a", "h3",
    )
    # Jeśli nadal brak – spróbuj tekstu linka do company
    if not name:
        try:
            els = card.find_elements(By.CSS_SELECTOR, "a[href*='alibaba.com']")
            for el in els:
                t = el.text.strip()
                if t and len(t) > 3 and "alibaba" not in t.lower():
                    name = t
                    break
        except Exception:
            pass

    if not name or len(name) < 2:
        return None

    # ── Link do profilu ──
    profile_url = safe_attr(
        "href",
        "a[href*='.alibaba.com/company']",
        "a[href*='company.html']",
        "[class*='company'] a",
        "a[href*='alibaba.com']",
    )
    if not profile_url:
        # zbuduj link z nazwy firmy jako fallback
        profile_url = source_url
    if profile_url and not profile_url.startswith("http"):
        profile_url = "https:" + profile_url

    # ── Ocena i liczba ocen ──
    # Format Alibaby: <div class="id-font-bold ...">4.9/5<span ...>(<span class="id-underline">132</span>)</span></div>
    full_text    = card.text
    rating       = 0.0
    review_count = ""

    # 1. Dedykowany selektor – div zawierający "X.X/5"
    rating_el = None
    try:
        rating_el = card.find_element(By.CSS_SELECTOR, "div[class*='id-font-bold']")
    except NoSuchElementException:
        pass

    if rating_el:
        raw = rating_el.text.strip()          # np. "4.9/5\n(132)"
        m = re.search(r"(\d+\.\d+)\s*/\s*5", raw)
        if m:
            try:
                rating = float(m.group(1))
            except ValueError:
                pass

    # 2. Liczba ocen – szukamy w HTML karty
    # Na stronie wyników: <span class="">48</span> (pusta klasa, sama liczba)
    # Fallback: span.id-underline
    if not review_count:
        try:
            card_html = card.get_attribute("outerHTML") or ""
            # Metoda A: span z pustą klasą zawierający samą liczbę
            m = re.search(r'<span[^>]*\bclass=""[^>]*>(\d+)</span>', card_html)
            if m and int(m.group(1)) > 0:
                review_count = m.group(1)
            # Metoda B: span.id-underline
            if not review_count:
                m = re.search(r'<span[^>]*\bclass="id-underline"[^>]*>(\d+)</span>', card_html)
                if m and int(m.group(1)) > 0:
                    review_count = m.group(1)
        except Exception:
            pass

    # 3. Fallback na tekst karty jeśli selektory nie trafiły
    if rating == 0.0:
        m = re.search(r"(\d+\.\d+)\s*/\s*5", full_text)
        if m:
            try:
                rating = float(m.group(1))
            except ValueError:
                pass

    # ── Lata na Alibaba ──
    years = safe_text(
        "[class*='year']", "[class*='gold-supplier']", "[class*='member-year']",
    )
    if not years:
        m = re.search(r"(\d+)\s*(yr|yrs|years?)\b", full_text, re.I)
        if m:
            years = m.group(0)

    # ── Lokalizacja ──
    location = safe_text(
        "[class*='country']", "[class*='location']",
        "[class*='province']", "[class*='city']",
    )
    if not location:
        m = re.search(
            r"(China|Guangdong|Zhejiang|Fujian|Jiangsu|Shandong|Hebei|"
            r"Guangzhou|Shenzhen|Yiwu|Dongguan|Foshan|Wenzhou|Hangzhou|"
            r"Beijing|Shanghai|Ningbo|Quanzhou)",
            full_text, re.I
        )
        if m:
            location = m.group(0)
    if not location:
        location = "Chiny"

    # ── Główne produkty ──
    products = safe_text(
        "[class*='main-product']", "[class*='products']", "[class*='goods']",
    )
    if not products:
        products = phrase

    # ── Czas odpowiedzi ──
    response = safe_text("[class*='response']", "[class*='reply']")

    # ── Weryfikacja dostawcy ──
    # Na stronie wyników: <img class="verified-supplier-icon" ...>
    # Szukamy w HTML karty (regex na outerHTML – szybkie, bez osobnych zapytań DOM)
    verified = "Nie"
    try:
        card_html_v = card.get_attribute("outerHTML") or ""
        if 'verified-supplier-icon' in card_html_v:
            verified = "Tak"
        elif "O1CN01I6gZX524hKekOpOzA" in card_html_v or "7422-55-tps-58-15" in card_html_v:
            verified = "Tak"
    except Exception:
        pass

    # Status zawsze startuje jako "Nowy" – pracownik sam decyduje dalej
    status = "Nowy"

    return {
        "Status":              status,
        "Fraza wyszukiwania":  phrase,
        "Nazwa dostawcy":      name,
        "Zweryfikowany":       verified,
        "Ocena (★)":           rating if rating > 0 else "–",
        "Liczba ocen":         review_count,
        "Lokalizacja":         location,
        "Lat na Alibaba":      years,
        "Główne produkty":     products[:200],
        "Czas odpowiedzi":     response,
        "Link do profilu":     profile_url or source_url,
    }


def enrich_from_profile(driver, supplier: dict) -> dict:
    """Odwiedza stronę produktu/profilu dostawcy i uzupełnia ocenę + weryfikację."""
    url = supplier.get("Link do profilu", "")
    if not url or not url.startswith("http"):
        return supplier

    try:
        driver.get(url)
    except Exception:
        pass   # page_load_strategy=none – ignorujemy wyjątki

    # Minimalne oczekiwanie – tylko tyle żeby strona zaczęła zwracać HTML
    time.sleep(1.0)

    # Pobierz page_source RAZ – szybka operacja, dalej szukamy regexem w Pythonie
    # (zamiast wolnych zapytań Selenium do każdego elementu DOM)
    try:
        src = driver.page_source
    except Exception:
        return supplier

    # Szybki check CAPTCHA w page source
    if any(signal in src.lower() for signal in CAPTCHA_SIGNALS):
        check_and_handle_captcha(driver)
        try:
            src = driver.page_source
        except Exception:
            return supplier

    # ── Weryfikacja – szukaj badge SVG w HTML (regex, błyskawiczne) ──
    if "O1CN01I6gZX524hKekOpOzA" in src or "7422-55-tps-58-15" in src:
        supplier["Zweryfikowany"] = "Tak"

    # ── Liczba ocen – <span class="id-underline">836</span> ──
    m = re.search(r'class="id-underline"[^>]*>(\d+)<', src)
    if m and int(m.group(1)) > 0:
        supplier["Liczba ocen"] = m.group(1)

    # ── Ocena – format "4.9/5" gdziekolwiek w HTML ──
    if str(supplier.get("Ocena (★)", "–")) in ("–", "", "0"):
        m = re.search(r"(\d+\.\d+)\s*/\s*5", src)
        if m:
            try:
                supplier["Ocena (★)"] = float(m.group(1))
            except ValueError:
                pass

    return supplier


def fallback_links(driver, phrase: str, source_url: str) -> list[dict]:
    """Zbiera linki do stron firmowych gdy brak kart."""
    results = []
    try:
        links = driver.find_elements(By.CSS_SELECTOR, "a[href*='.alibaba.com']")
        seen = set()
        for link in links:
            href = link.get_attribute("href") or ""
            if "company" in href.lower() and href not in seen:
                seen.add(href)
                name = link.text.strip() or href.split("/")[-1]
                if name and len(name) > 2:
                    results.append({
                        "Fraza wyszukiwania": phrase,
                        "Nazwa dostawcy":     name,
                        "Ocena (★)":          "–",
                        "Zweryfikowany":      "Nie",
                        "Ocena (★)":          "–",
                        "Liczba ocen":        "",
                        "Lokalizacja":        "Chiny",
                        "Lat na Alibaba":     "",
                        "Główne produkty":    phrase,
                        "Czas odpowiedzi":    "",
                        "Link do profilu":    href,
                        "Status":             "Nowy",
                    })
    except Exception as e:
        print(f"    Fallback error: {e}")
    return results


# ─────────────────────────────────────────────
#  FILTROWANIE – min. X gwiazdek
# ─────────────────────────────────────────────
def filter_by_rating(suppliers: list[dict], min_stars: float) -> list[dict]:
    if min_stars <= 0:
        return suppliers
    filtered = []
    for s in suppliers:
        try:
            if float(s.get("Ocena (★)", 0)) >= min_stars:
                filtered.append(s)
        except (TypeError, ValueError):
            pass  # brak oceny – pomijamy
    return filtered


# ─────────────────────────────────────────────
#  ZAPIS DO EXCELA
# ─────────────────────────────────────────────
def save_to_excel(all_suppliers: list[dict], filename: str):
    wb = openpyxl.Workbook()

    ws_all = wb.active
    ws_all.title = "Wszyscy dostawcy"
    _write_sheet(ws_all, all_suppliers)

    phrases = list(dict.fromkeys(s["Fraza wyszukiwania"] for s in all_suppliers))
    for phrase in phrases:
        ws = wb.create_sheet(title=phrase[:25].strip())
        _write_sheet(ws, [s for s in all_suppliers if s["Fraza wyszukiwania"] == phrase])

    ws_help = wb.create_sheet(title="📖 Instrukcja")
    _write_instructions(ws_help)

    wb.save(filename)
    print(f"\n✅  Plik zapisany: {filename}")


def _write_sheet(ws, rows: list[dict]):
    if not rows:
        ws.append(["Brak wyników dla tej frazy"])
        return

    # ── Sortowanie według priorytetu statusu ──
    priority = {
        "Zaakceptowany":            0,
        "W trakcie":                1,
        "Sprawdzony":               2,
        "Nowy":                     3,
        "Do ponownego sprawdzenia": 4,
        "Odrzucony":                5,
    }
    rows = sorted(rows, key=lambda r: priority.get(r.get("Status", "Nowy"), 3))

    columns = list(rows[0].keys())
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor="1B4F72")
    h_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Nagłówek ──
    ws.append(columns)
    for ci, cn in enumerate(columns, 1):
        c = ws.cell(1, ci)
        c.fill = h_fill; c.font = h_font; c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 28

    # ── Dane ──
    last_row = len(rows) + 1
    status_col_idx    = columns.index("Status") + 1 if "Status" in columns else None
    status_col_letter = get_column_letter(status_col_idx) if status_col_idx else "A"

    # Wiersze startują BEZ koloru – kolor pojawia się gdy pracownik
    # zmieni Status w dropdownie (obsługuje to Conditional Formatting poniżej)
    for ri, row in enumerate(rows, 2):
        for ci, cn in enumerate(columns, 1):
            val  = row.get(cn, "")
            cell = ws.cell(ri, ci, val)
            cell.border    = border
            cell.alignment = Alignment(vertical="center")

            if cn == "Status":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font      = Font(bold=False, color="444444")
            elif cn == "Link do profilu" and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font      = Font(color="1F618D", underline="single")
                cell.value     = "🔗 Otwórz profil"
            elif cn == "Ocena (★)":
                try:
                    cell.value = float(val)
                    cell.number_format = "0.0"
                except (TypeError, ValueError):
                    pass
            elif cn == "Zweryfikowany":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(
                    color="1A7F37" if val == "Tak" else "767676",
                    bold=(val == "Tak")
                )
        ws.row_dimensions[ri].height = 20

    # ── Conditional Formatting – kolory pojawiają się gdy pracownik zmieni status ──
    # Nowy = brak koloru (domyślny biały), pozostałe kolorowane
    data_range = f"A2:{get_column_letter(len(columns))}{last_row}"
    ref = f"${status_col_letter}2"
    CF_RULES = {
        "W trakcie":                "FFEB9C",   # żółty
        "Sprawdzony":               "BDD7EE",   # niebieski
        "Zaakceptowany":            "C6EFCE",   # zielony
        "Odrzucony":                "FFC7CE",   # czerwony
        "Do ponownego sprawdzenia": "FFD966",   # pomarańczowy
    }
    for status_val, color in CF_RULES.items():
        ws.conditional_formatting.add(data_range,
            FormulaRule(formula=[f'{ref}="{status_val}"'],
                        fill=PatternFill("solid", fgColor=color)))

    # ── Dropdown w kolumnie Status ──
    if status_col_idx:
        dv = DataValidation(
            type="list",
            formula1='"Nowy,W trakcie,Sprawdzony,Zaakceptowany,Odrzucony,Do ponownego sprawdzenia"',
            allow_blank=False,
            showDropDown=False,
        )
        dv.sqref = f"{status_col_letter}2:{status_col_letter}{last_row}"
        ws.add_data_validation(dv)

    # ── Szerokości kolumn ──
    col_widths = {
        "Status": 16, "Fraza wyszukiwania": 24, "Nazwa dostawcy": 38,
        "Zweryfikowany": 14, "Ocena (★)": 10, "Liczba ocen": 12,
        "Lokalizacja": 24, "Lat na Alibaba": 14,
        "Główne produkty": 42, "Czas odpowiedzi": 18, "Link do profilu": 20,
    }
    for ci, cn in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(cn, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _format_sheets(sh, ws, num_cols: int, last_row: int):
    """Stosuje pełne formatowanie arkusza Google Sheets:
    nagłówek, szerokości kolumn, dropdown statusu, conditional formatting."""

    sheet_id = ws.id
    max_rows = max(last_row + 100, 500)

    # Kolory statusów (RGB 0–1)
    STATUS_COLORS = {
        "W trakcie":                {"red": 1.000, "green": 0.922, "blue": 0.612},  # żółty
        "Sprawdzony":               {"red": 0.741, "green": 0.843, "blue": 0.933},  # niebieski
        "Zaakceptowany":            {"red": 0.776, "green": 0.937, "blue": 0.808},  # zielony
        "Odrzucony":                {"red": 1.000, "green": 0.780, "blue": 0.808},  # czerwony
        "Do ponownego sprawdzenia": {"red": 1.000, "green": 0.851, "blue": 0.400},  # pomarańczowy
        # Nowy = bez koloru
    }

    requests = []

    # 0. Format kolumny Ocena (★) = E (indeks 4) jako liczba "0.0"
    # Bez tego Google Sheets może zinterpretować float 4.9 jako datę 9 kwietnia
    requests.append({"repeatCell": {
        "range": {
            "sheetId": sheet_id,
            "startRowIndex": 1,
            "endRowIndex": max_rows,
            "startColumnIndex": 4,   # kolumna E = Ocena (★)
            "endColumnIndex": 5,
        },
        "cell": {"userEnteredFormat": {
            "numberFormat": {"type": "NUMBER", "pattern": "0.0"},
        }},
        "fields": "userEnteredFormat.numberFormat",
    }})

    # 1. Wyrównanie tekstu – centrowanie wszystkich komórek danych
    requests.append({"repeatCell": {
        "range": {
            "sheetId": sheet_id,
            "startRowIndex": 1,
            "endRowIndex": max_rows,
            "startColumnIndex": 0,
            "endColumnIndex": num_cols,
        },
        "cell": {"userEnteredFormat": {
            "horizontalAlignment": "CENTER",
            "verticalAlignment": "MIDDLE",
            "wrapStrategy": "WRAP",
        }},
        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,wrapStrategy)",
    }})

    # 2. Szerokości kolumn (px)
    widths = [120, 170, 250, 90, 70, 80, 150, 95, 270, 130, 130]
    for i, w in enumerate(widths[:num_cols]):
        requests.append({"updateDimensionProperties": {
            "range": {"sheetId": sheet_id, "dimension": "COLUMNS",
                      "startIndex": i, "endIndex": i + 1},
            "properties": {"pixelSize": w},
            "fields": "pixelSize",
        }})

    # 2a. Wysokość wiersza nagłówka
    requests.append({"updateDimensionProperties": {
        "range": {"sheetId": sheet_id, "dimension": "ROWS",
                  "startIndex": 0, "endIndex": 1},
        "properties": {"pixelSize": 36},
        "fields": "pixelSize",
    }})
    # 2b. Jednakowa wysokość wszystkich wierszy danych (od wiersza 2)
    requests.append({"updateDimensionProperties": {
        "range": {"sheetId": sheet_id, "dimension": "ROWS",
                  "startIndex": 1, "endIndex": max_rows},
        "properties": {"pixelSize": 40},
        "fields": "pixelSize",
    }})

    # 3. Zamróź wiersz 1
    requests.append({"updateSheetProperties": {
        "properties": {"sheetId": sheet_id,
                       "gridProperties": {"frozenRowCount": 1}},
        "fields": "gridProperties.frozenRowCount",
    }})

    # 4. Dropdown statusu w kolumnie A (dane od wiersza 2)
    requests.append({"setDataValidation": {
        "range": {"sheetId": sheet_id,
                  "startRowIndex": 1, "endRowIndex": max_rows,
                  "startColumnIndex": 0, "endColumnIndex": 1},
        "rule": {
            "condition": {
                "type": "ONE_OF_LIST",
                "values": [
                    {"userEnteredValue": "Nowy"},
                    {"userEnteredValue": "W trakcie"},
                    {"userEnteredValue": "Sprawdzony"},
                    {"userEnteredValue": "Zaakceptowany"},
                    {"userEnteredValue": "Odrzucony"},
                    {"userEnteredValue": "Do ponownego sprawdzenia"},
                ],
            },
            "showCustomUi": True,
            "strict": False,
        },
    }})

    # 5. Conditional formatting – koloruje cały wiersz na podstawie statusu w kol. A
    cf_range = {"sheetId": sheet_id,
                "startRowIndex": 1, "endRowIndex": max_rows,
                "startColumnIndex": 0, "endColumnIndex": num_cols}
    for status_val, color in STATUS_COLORS.items():
        requests.append({"addConditionalFormatRule": {
            "rule": {
                "ranges": [cf_range],
                "booleanRule": {
                    "condition": {
                        "type": "CUSTOM_FORMULA",
                        "values": [{"userEnteredValue": f'=$A2="{status_val}"'}],
                    },
                    "format": {"backgroundColor": color},
                },
            },
            "index": 0,
        }})

    try:
        sh.batch_update({"requests": requests})
    except Exception as e:
        print(f"   ⚠️  Formatowanie Sheets (niekrytyczne): {e}")


def save_to_sheets(suppliers: list[dict]):
    """Zapisuje dostawców do Google Sheets, pomijając duplikaty."""

    if not GOOGLE_SHEETS_ENABLED:
        return

    if not GSPREAD_AVAILABLE:
        print("\n⚠️  Brak bibliotek Google Sheets.")
        print("   Zainstaluj: pip install gspread google-auth --break-system-packages")
        return

    import os
    if not os.path.exists(GOOGLE_CREDS_FILE):
        print(f"\n⚠️  Nie znaleziono pliku z kluczem Google: '{GOOGLE_CREDS_FILE}'")
        print("   Wykonaj kroki SETUP opisane na górze skryptu.")
        return

    print("\n📊 Łączę z Google Sheets...")
    try:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_file(GOOGLE_CREDS_FILE, scopes=scopes)
        gc    = gspread.authorize(creds)
    except Exception as e:
        print(f"   ❌ Błąd autoryzacji Google: {e}")
        return

    HEADERS = [
        "Status", "Fraza wyszukiwania", "Nazwa dostawcy", "Zweryfikowany",
        "Ocena (★)", "Liczba ocen", "Lokalizacja", "Lat na Alibaba",
        "Główne produkty", "Czas odpowiedzi", "Link do profilu",
    ]
    NUM_COLS = len(HEADERS)

    # Otwórz lub stwórz arkusz
    try:
        sh = gc.open(GOOGLE_SHEET_NAME)
        print(f"   ✓ Połączono z arkuszem: '{GOOGLE_SHEET_NAME}'")
    except gspread.SpreadsheetNotFound:
        sh = gc.create(GOOGLE_SHEET_NAME)
        print(f"   ✓ Stworzono nowy arkusz: '{GOOGLE_SHEET_NAME}'")
        sh.share(None, perm_type="anyone", role="reader")

    ws = sh.sheet1
    existing_data = ws.get_all_values()

    # ── Sprawdź czy wiersz 1 to nagłówki, jeśli nie – wstaw ──
    has_headers = (
        len(existing_data) > 0
        and len(existing_data[0]) > 0
        and existing_data[0][0].strip() == "Status"
    )
    if not has_headers:
        if not existing_data:
            ws.append_row(HEADERS)
        else:
            ws.insert_row(HEADERS, index=1)   # wstaw nad istniejącymi danymi
            existing_data = [HEADERS] + existing_data
        # Sformatuj nagłówek
        ws.format("A1:K1", {
            "backgroundColor": {"red": 0.11, "green": 0.31, "blue": 0.45},
            "textFormat": {
                "bold": True,
                "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                "fontSize": 11,
            },
            "horizontalAlignment": "CENTER",
            "verticalAlignment": "MIDDLE",
        })

    # ── Zbierz istniejące klucze (nazwa + fraza) do deduplicacji ──
    existing_names = set()
    for row in existing_data[1:]:
        if len(row) >= 3:
            key = (row[2].lower().strip(), row[1].lower().strip())  # Nazwa, Fraza
            existing_names.add(key)

    # ── Filtruj duplikaty ──
    new_suppliers = []
    for s in suppliers:
        key = (
            s.get("Nazwa dostawcy", "").lower().strip(),
            s.get("Fraza wyszukiwania", "").lower().strip(),
        )
        if key not in existing_names:
            new_suppliers.append(s)
            existing_names.add(key)

    print(f"   Sprawdzono {len(suppliers)} dostawców → {len(new_suppliers)} nowych (bez duplikatów)")

    if not new_suppliers:
        print("   ℹ️  Wszyscy dostawcy już są w arkuszu.")
        print(f"   🔗 Link: {sh.url}")
        _format_sheets(sh, ws, NUM_COLS, len(existing_data))
        return

    # ── Posortuj według priorytetu przed wgraniem ──
    prio_map = {
        "Zaakceptowany": 1, "W trakcie": 2, "Sprawdzony": 3,
        "Nowy": 4, "Do ponownego sprawdzenia": 5, "Odrzucony": 6,
    }
    new_suppliers = sorted(new_suppliers,
                           key=lambda s: prio_map.get(s.get("Status", "Nowy"), 4))

    # ── Przygotuj wiersze (URL jako zwykły tekst – Sheets sam robi klikalny link) ──
    rows_to_add = []
    for s in new_suppliers:
        row = []
        for h in HEADERS:
            val = s.get(h, "")
            if h == "Ocena (★)":
                # Wysyłaj jako float, nie string – "4.9" jako tekst Google Sheets
                # interpretuje jako datę 9 kwietnia (4/9 → 2026-04-09)
                try:
                    row.append(float(val) if val and val != "–" else "")
                except (ValueError, TypeError):
                    row.append("")
            else:
                row.append(str(val))
        rows_to_add.append(row)

    # ── Dodaj partiami (limit API: ~80 wierszy naraz) ──
    BATCH = 80
    for i in range(0, len(rows_to_add), BATCH):
        ws.append_rows(rows_to_add[i:i+BATCH], value_input_option="USER_ENTERED")
        if len(rows_to_add) > BATCH:
            time.sleep(1)

    last_row = len(existing_data) + len(new_suppliers)
    _format_sheets(sh, ws, NUM_COLS, last_row)

    print(f"   ✅ Dodano {len(new_suppliers)} nowych wierszy!")
    print(f"   🔗 Link do arkusza: {sh.url}")


def _write_instructions(ws):
    ws.column_dimensions["A"].width = 85
    ws.cell(1, 1, "📖  INSTRUKCJA – Wyszukiwarka dostawców Alibaba").font = Font(bold=True, size=14, color="1B4F72")
    ws.row_dimensions[1].height = 30
    lines = [
        "",
        "JAK UŻYWAĆ TEGO PLIKU:",
        "1. Arkusz 'Wszyscy dostawcy' zawiera wszystkich znalezionych dostawców (≥4★).",
        "2. Każda zakładka to jedna fraza wyszukiwania.",
        "3. Kliknij '🔗 Otwórz profil' aby przejść do dostawcy na Alibaba.",
        "4. Użyj filtrów (▼ w nagłówku) aby sortować po ocenie, lokalizacji itp.",
        "",
        "KONTAKT Z DOSTAWCĄ:",
        "a) Kliknij 'Otwórz profil' → strona firmy na Alibaba.",
        "b) Kliknij 'Contact Supplier' na stronie Alibaby.",
        "c) Opisz swoje potrzeby: rodzaj ramek, ilości, materiały.",
        "d) Poproś o próbki (samples) i cennik (MOQ / price list).",
        "",
        "WSKAZÓWKI:",
        "• 'Verified Supplier' lub 'Gold Supplier' = dodatkowa weryfikacja.",
        "• Response Rate >80% i ≤24h = rzetelny dostawca.",
        "• Wybieraj dostawców z >5 lat na platformie.",
        "• Trade Assurance = ochrona płatności przez Alibabę.",
        "",
        "JAK PONOWNIE URUCHOMIĆ:",
        "  python alibaba_dostawcy.py",
        "",
        f"Data ostatniego uruchomienia: {date.today()}",
    ]
    for i, line in enumerate(lines, 2):
        cell = ws.cell(i, 1, line)
        cell.alignment = Alignment(wrap_text=True)
        if any(line.startswith(x) for x in ["JAK", "KONTAKT", "WSKAZÓWKI"]):
            cell.font = Font(bold=True, color="1B4F72", size=11)
        ws.row_dimensions[i].height = 18


# ─────────────────────────────────────────────
#  POMOCNIK TIMERA
# ─────────────────────────────────────────────
def fmt_time(seconds: float) -> str:
    """Formatuje sekundy jako MM:SS."""
    m = int(seconds) // 60
    s = int(seconds) % 60
    return f"{m:02d}:{s:02d}"


def print_timer(start_ts: float, phrase_idx: int, total_phrases: int, found: int):
    """Drukuje linię statusu z timerem, postępem i liczbą znalezionych."""
    elapsed   = time.time() - start_ts
    remaining = max(0, MAX_RUNTIME_MIN * 60 - elapsed)
    bar_done  = int(20 * phrase_idx / total_phrases)
    bar       = "█" * bar_done + "░" * (20 - bar_done)
    print(
        f"\n  ⏱  [{fmt_time(elapsed)} upłynęło | ~{fmt_time(remaining)} pozostało | "
        f"{phrase_idx}/{total_phrases} fraz]  [{bar}]  "
        f"Dostawców: {found}"
    )


# ─────────────────────────────────────────────
#  GŁÓWNA PĘTLA
# ─────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  ALIBABA SUPPLIER FINDER v3.0 – undetected-chromedriver")
    print(f"  Frazy: {len(SEARCH_PHRASES)}  |  Min. ocena: {MIN_STARS}★  |  Stron/frazę: {MAX_PAGES}")
    print(f"  Limit czasu: {MAX_RUNTIME_MIN} min  |  CAPTCHA timeout: {CAPTCHA_WAIT_S}s")
    print(f"  Przeglądarka: widoczna (tryb anty-bot)")
    print("=" * 60)

    print("\n🚀 Uruchamiam przeglądarkę Chrome (undetected)...")
    try:
        driver = create_driver()
    except Exception as e:
        print(f"\n❌  Nie można uruchomić Chrome: {e}")
        print("\nSprawdź czy Google Chrome jest zainstalowany.")
        print("Pobierz z: https://www.google.com/chrome/")
        sys.exit(1)

    all_suppliers = []
    start_time    = time.time()
    pages_since_restart = 0
    RESTART_EVERY = 20   # restart Chrome co tyle stron – czyści pamięć przeglądarki

    def time_left_ok() -> bool:
        """True jeśli mamy jeszcze czas (< MAX_RUNTIME_MIN)."""
        elapsed_min = (time.time() - start_time) / 60
        return elapsed_min < MAX_RUNTIME_MIN

    try:
        for pi, phrase in enumerate(SEARCH_PHRASES, 1):

            # ── Sprawdź limit czasu przed każdą frazą ──
            if not time_left_ok():
                print(f"\n⏰  Limit {MAX_RUNTIME_MIN} minut osiągnięty – kończę zbieranie.")
                break

            print_timer(start_time, pi - 1, len(SEARCH_PHRASES), len(all_suppliers))
            print(f"\n[{pi}/{len(SEARCH_PHRASES)}] Szukam: \"{phrase}\"")
            phrase_results = []

            for page in range(1, MAX_PAGES + 1):

                # ── Sprawdź limit czasu przed każdą stroną ──
                if not time_left_ok():
                    print(f"\n⏰  Limit czasu – kończę w trakcie frazy \"{phrase}\".")
                    break

                elapsed_s  = time.time() - start_time
                remaining_s = max(0, MAX_RUNTIME_MIN * 60 - elapsed_s)
                print(
                    f"  → Strona {page}/{MAX_PAGES}  "
                    f"[⏱ {fmt_time(elapsed_s)} | ~{fmt_time(remaining_s)} pozostało]",
                    end=" ", flush=True
                )

                try:
                    results = search_suppliers(driver, phrase, page)
                except Exception as e:
                    print(f"\n  ⚠️  Błąd przeglądarki ({type(e).__name__}) – restartuję Chrome...")
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    time.sleep(5)
                    driver = create_driver()
                    pages_since_restart = 0
                    print("  ✓  Przeglądarka zrestartowana – pomijam tę stronę.\n")
                    results = []

                filtered = filter_by_rating(results, MIN_STARS)
                phrase_results.extend(filtered)
                print(f"wyników: {len(results)}, po filtrze: {len(filtered)}")
                pages_since_restart += 1

                # ── Restart Chrome co RESTART_EVERY stron – czyści pamięć ──
                if pages_since_restart >= RESTART_EVERY and time_left_ok():
                    print(f"\n  🔄  Restartuję przeglądarkę (odświeżam pamięć po {RESTART_EVERY} stronach)...")
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    time.sleep(3)
                    driver = create_driver()
                    pages_since_restart = 0
                    print("  ✓  Przeglądarka zrestartowana.\n")

                if page < MAX_PAGES and time_left_ok():
                    time.sleep(random.uniform(*DELAY_PAGE))

            # Usuń duplikaty między frazami
            seen   = {s["Nazwa dostawcy"] for s in all_suppliers}
            unique = [s for s in phrase_results if s["Nazwa dostawcy"] not in seen]
            all_suppliers.extend(unique)
            print(f"  ✔  Nowych unikalnych dostawców z tej frazy: {len(unique)}")

            # ── Zapis po każdej frazie – nie czekamy na koniec ──
            if unique:
                print(f"  💾  Zapisuję {len(unique)} nowych dostawców...")
                save_to_excel(all_suppliers, OUTPUT_FILE)
                save_to_sheets(unique)   # tylko nowi z tej frazy (Sheets samo deduplikuje)
                print(f"  ✅  Łącznie w bazie: {len(all_suppliers)} dostawców")

            if pi < len(SEARCH_PHRASES) and time_left_ok():
                delay = random.uniform(*DELAY_PHRASE)
                print(f"  ⏳ Czekam {delay:.0f}s przed następną frazą...")
                time.sleep(delay)

        # ══════════════════════════════════════════════════════
        #  FAZA 2 – Wzbogacanie: odwiedź profil każdego dostawcy
        # ══════════════════════════════════════════════════════
        if all_suppliers and ENRICH_PROFILES and time_left_ok():
            total_e = len(all_suppliers)
            print(f"\n{'═'*60}")
            print(f"  🔍  FAZA 2: Sprawdzam profile dostawców ({total_e} unikalnych)...")
            print(f"{'═'*60}")

            for i, s in enumerate(all_suppliers):
                if not time_left_ok():
                    print(f"\n⏰  Limit czasu – wzbogacono {i}/{total_e} profili.")
                    break

                name = s.get("Nazwa dostawcy", "?")[:38]
                elapsed_s   = time.time() - start_time
                remaining_s = max(0, MAX_RUNTIME_MIN * 60 - elapsed_s)
                print(
                    f"  [{i+1:>3}/{total_e}] {name:<38} "
                    f"[⏱ {fmt_time(elapsed_s)} | ~{fmt_time(remaining_s)} pozostało]",
                    end=" ", flush=True
                )

                all_suppliers[i] = enrich_from_profile(driver, s)

                v  = all_suppliers[i].get("Zweryfikowany", "Nie")
                r  = all_suppliers[i].get("Ocena (★)", "–")
                lc = all_suppliers[i].get("Liczba ocen", "–")
                print(f"→ Weryfikacja: {v} | Ocena: {r} | Ocen: {lc}")

                if i < total_e - 1 and time_left_ok():
                    time.sleep(random.uniform(*DELAY_ENRICH))

    except KeyboardInterrupt:
        print("\n\n⚠️  Przerwano przez użytkownika – zapisuję co mam...")

    finally:
        try:
            driver.quit()
        except Exception:
            pass
        print("\n🔒 Przeglądarka zamknięta.")

    # ── Podsumowanie ──
    total_elapsed = time.time() - start_time
    print(f"\n{'=' * 60}")
    print(f"  ⏱  Czas działania: {fmt_time(total_elapsed)}")
    total = len(all_suppliers)
    print(f"  Łącznie: {total} unikalnych dostawców (≥{MIN_STARS}★)")

    if total > 0:
        # Końcowy zapis – aktualizuje Excel z wszystkimi danymi (Sheets już ma wszystko na bieżąco)
        save_to_excel(all_suppliers, OUTPUT_FILE)
        print(f"  📂 Lokalny Excel: {OUTPUT_FILE}")
        print(f"  📊 Google Sheets: dane wysyłane na bieżąco po każdej frazie")
    else:
        print("\n  ⚠️  Nie znaleziono dostawców.")
        print("  Możliwe przyczyny:")
        print("  1. Alibaba pokazała CAPTCHA – spróbuj ponownie")
        print("  2. Alibaba zmieniła selektory HTML – zgłoś do autora skryptu.")
        print("  3. Problem z połączeniem internetowym.")

    print("=" * 60)


if __name__ == "__main__":
    main()
